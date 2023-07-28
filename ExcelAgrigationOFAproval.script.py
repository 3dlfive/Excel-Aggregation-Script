import os
import datetime
import openpyxl 

from openpyxl import Workbook

currentDT = datetime.datetime.now()
print ("Current Microsecond is: %d" % currentDT.microsecond)
UID = str(currentDT.microsecond)


ROOT_DIR = os.path.dirname(os.path.abspath(__file__))  + "\excelFiles"
# Создание ексель файла
def initNewFile():
    wb = Workbook()
    sheet = wb.worksheets[0]
    # Хедер таблицы
    new_row = ["FileName","Product code","Intiated By","Kateryna Khriienko","Alla Austin","Max Zhukov","Mikey Garcia","Ekaterina Bugrova","Greg Campbell","PN"]
    sheet.append(new_row)
    wb.save("Result" + UID + ".xlsx")
#Запись данных в ексель файл
def addDataToExcelFile(filename,wb1value1,wb1value2,wb2value1,wb2value2,wb2value3,wb2value4,wb2value5,wb2value6,wb2value7):
    excel_filename = openpyxl.load_workbook(os.path.dirname(os.path.abspath(__file__)) + "\\Result" + UID + ".xlsx")
    wb = excel_filename
    sheet = wb.worksheets[0]
    new_row = [filename,wb1value1,wb1value2,wb2value1,wb2value2,wb2value3,wb2value4,wb2value5,wb2value6,wb2value7]
    sheet.append(new_row)
    wb.save("Result" + UID + ".xlsx")

#получение данных
def extractDataFromExcelFile(file):
    excel_readFile = openpyxl.load_workbook(os.path.dirname(os.path.abspath(__file__)) + "\excelFiles\\" + file)
    wb2 = excel_readFile
    print(wb2.sheetnames)
    sheet2 = wb2['Checklist']
    prdCode = sheet2["B10"].value
    initiatedBy = sheet2["B2"].value
    sheet3 = wb2['Approval Log']
    aprovel01 = sheet3["C2"].value
    aprovel02 = sheet3["C3"].value
    aprovel03 = sheet3["C4"].value
    aprovel04 = sheet3["C5"].value
    aprovel05 = sheet3["C6"].value
    aprovel06 = sheet3["C7"].value
    aprovel07 = sheet3["D10"].value
    addDataToExcelFile(file,prdCode,initiatedBy,aprovel01,aprovel02,aprovel03,aprovel04,aprovel05,aprovel06,aprovel07)
   
    print("Step of extracting and saving data is finished for one file")
    #Add data to file Excel

# Сбор названий файлов
def scanningFileNames(dirPath):
    list_f =os.listdir(dirPath)
    for file in list_f:
        if file.endswith('.xlsx'):
            extractDataFromExcelFile(file)
            print("File name done: ", "\n",file)
initNewFile()
scanningFileNames(ROOT_DIR)